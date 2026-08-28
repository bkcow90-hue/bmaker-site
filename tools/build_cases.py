#!/usr/bin/env python3
"""실행 사례 원장 빌드 — data/cases.xlsx(원장 시트) 하나에서 사이트의 모든 원장 표면을 재생성한다.

생성/갱신 대상: cases.html · data/cases.csv · llms-full.txt(원장 섹션) · llms.txt(요약 한 줄) · sitemap.xml(/cases lastmod)
실행: python tools/build_cases.py   (저장소 루트 어디서든 가능)
실패 시: 어떤 행·열이 문제인지 한국어로 출력하고 아무 파일도 쓰지 않는다.
"""
import csv, json, re, sys, datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
FEE = "착수금·진행비 등 실행 전 비용은 일절 받지 않고, 자금이 실제 실행된 경우에만 성공보수를 받습니다."
COLS = ['사례ID','실행 연월','기관','자금명','실행 금액(만원)','금리','상환 조건','지역(시도)','업종','사업 형태','업력(년)',
        '신용점수 구간','연매출 구간','폐업 이력','체납 이력','기존 정책자금','동시 진행 자금','소요일','한 줄 메모','증빙 파일','사이트 공개']
REQ = ['사례ID','실행 연월','기관','자금명','실행 금액(만원)','지역(시도)','사이트 공개']

def die(msg):
    print(f"[원장 빌드 실패] {msg}"); sys.exit(1)

def norm_rate(v):
    """금리 표기 통일: 앞머리의 숫자를 '연 N.NN%' 틀로 맞춘다. 숫자와 뒤따르는 설명(괄호 등)은 그대로 보존.
    예: '3.65'→'연 3.65%' · '연 3.43(충북 육성자금)'→'연 3.43% (충북 육성자금)' · 숫자로 시작하지 않으면 입력 그대로."""
    v = str(v or "").strip()
    if not v: return ""
    m = re.match(r"^연?\s*(\d+(?:\.\d+)?)\s*%?\s*(.*)$", v)
    if not m: return v
    num, rest = m.group(1), m.group(2).strip()
    return f"연 {num}%" + (f" {rest}" if rest else "")

def _norm_header(h):
    return re.sub(r"\s+", " ", str(h or "")).strip()

def _rows_from_records(records, where):
    rows, ids = [], set()
    for i, d in records:
        if not any(v for v in d.values()): continue
        for c in REQ:
            if not d.get(c, ""): die(f"{where} {i}행: 필수 열 '{c}' 이 비어 있습니다.")
        if d["사례ID"] in ids: die(f"{where} {i}행: 사례ID '{d['사례ID']}' 중복입니다.")
        ids.add(d["사례ID"])
        if not re.match(r"^\d{4}-\d{2}$", d["실행 연월"]): die(f"{where} {i}행 실행 연월 '{d['실행 연월']}' — YYYY-MM 형식이어야 합니다.")
        try: d["amt"] = int(float(d["실행 금액(만원)"].replace(",", "")))
        except ValueError: die(f"{where} {i}행 실행 금액 '{d['실행 금액(만원)']}' — 만원 단위 숫자만.")
        if d["사이트 공개"].upper() not in ("Y", "N"): die(f"{where} {i}행 사이트 공개는 Y 또는 N.")
        joined = " ".join(v for v in d.values() if isinstance(v, str))
        if "갚" in joined: die(f"{where} {i}행: '갚다'류 표현 금지 — 상환/돌려주다 로 바꿔주세요.")
        if re.search(r"보장", joined): die(f"{where} {i}행: '보장' 표현 금지.")
        if re.search(r"수수료|성공보수.{0,12}%|\d+\s*%\s*(수수료|보수)", joined): die(f"{where} {i}행: 보수·수수료 관련 표기는 원장에 쓸 수 없습니다.")
        if d["증빙 파일"] and not (ROOT/"assets"/"cases"/d["증빙 파일"]).exists():
            die(f"{where} {i}행 증빙 파일 assets/cases/{d['증빙 파일']} 이 없습니다. 파일을 먼저 넣어주세요.")
        d["금리"] = norm_rate(d.get("금리", ""))
        rows.append(d)
    if not rows: die(f"{where}: 데이터 행이 없습니다.")
    return [d for d in rows if d["사이트 공개"].upper() == "Y"]

def load_rows():
    src = ROOT/"data"/"cases.source.csv"
    if src.exists():
        with open(src, encoding="utf-8-sig", newline="") as f:
            rdr = csv.reader(f)
            header = [_norm_header(h) for h in next(rdr)]
            missing = [c for c in REQ if c not in header]
            if missing: die(f"cases.source.csv 헤더에 {missing} 열이 없습니다. 시트 1행을 확인해주세요.")
            records = []
            for i, r in enumerate(rdr, start=2):
                vals = ["" if v is None else str(v).strip() for v in r] + [""] * len(header)
                d = {c: "" for c in COLS}
                for j, hname in enumerate(header):
                    if hname in COLS: d[hname] = vals[j]
                records.append((i, d))
        print(f"[원본] data/cases.source.csv ({len(records)}행)")
        return _rows_from_records(records, "cases.source.csv")
    wb = load_workbook(ROOT/"data"/"cases.xlsx", data_only=True)
    if "원장" not in wb.sheetnames: die("data/cases.xlsx 에 '원장' 시트가 없습니다.")
    ws = wb["원장"]
    records = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = ["" if v is None else str(v).strip() for v in (list(r)+[""]*21)[:21]]
        records.append((i, dict(zip(COLS, vals))))
    print("[원본] data/cases.xlsx")
    return _rows_from_records(records, "cases.xlsx")

def won2(m):
    e, man = divmod(int(m), 10000)
    return ((f"{e}억" if e else "") + ((" " if e and man else "") + f"{man:,}만" if man else "")) + "원"

def esc(s): return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

def agef(v):
    if not v: return ""
    try: a=float(v)
    except ValueError: return v
    return "1년 미만" if a<1 else f"{round(a)}년"

def build():
    D = load_rows()
    D.sort(key=lambda d:(d['실행 연월'], d['사례ID']), reverse=True)
    N=len(D); total=sum(d['amt'] for d in D)
    amts=sorted(d['amt'] for d in D); med=amts[N//2] if N%2 else (amts[N//2-1]+amts[N//2])//2
    days=sorted(int(float(d['소요일'])) for d in D if d['소요일'])
    dmed=(days[len(days)//2] if len(days)%2 else (days[len(days)//2-1]+days[len(days)//2])//2) if days else 0
    nc=sum(1 for d in D if d['폐업 이력']=='있음'); nt=sum(1 for d in D if d['체납 이력']=='있음')
    np_=sum(1 for d in D if d['기존 정책자금']=='있음'); nl=sum(1 for d in D if d['신용점수 구간'] in ('600점대','500점대 이하'))
    nev=sum(1 for d in D if d['증빙 파일'])
    yms=sorted(d['실행 연월'] for d in D); y0,m0=yms[0].split('-'); y1,m1=yms[-1].split('-')
    today=datetime.date.today()
    T={'N':N,'TOTAL_KR':won2(total),'MED_KR':won2(med),'RANGE_KR':f"{won2(amts[0]).replace('원','')}~{won2(amts[-1])}",
       'DMED':dmed,'DN':len(days),'NEV':nev,'n_closed':nc,'n_tax':nt,'n_prior':np_,'n_low':nl,
       'PERIOD_KR':f"{int(y0)}년 {int(m0)}월~{int(y1)}년 {int(m1)}월",'PERIOD_LONG':f"{int(y0)}년 {int(m0)}월부터 {int(y1)}년 {int(m1)}월까지",
       'PERIOD_SHORT':f"{y0}.{m0}~{y1}.{m1}",'ASOF_KR':f"{today.year}년 {today.month}월 {today.day}일"}
    def marks(d): return " · ".join(l for k,l in (('폐업 이력','폐업'),('체납 이력','체납'),('기존 정책자금','기존대출')) if d[k]=='있음') or "—"
    rows_html="".join(
      f'<tr id="row-{d["사례ID"]}"><td>{d["실행 연월"]}</td><td>{esc(d["기관"])}</td><td>{esc(d["자금명"])}</td>'
      f'<td class="num">{d["amt"]:,}</td><td>{esc(d["금리"]) or "—"}</td><td>{esc(d["상환 조건"]) or "—"}</td><td>{esc(d["지역(시도)"])}</td>'
      f'<td>{esc(d["업종"]) or "—"}</td><td>{esc(d["사업 형태"]) or "—"}</td><td>{agef(d["업력(년)"]) or "—"}</td><td>{d["신용점수 구간"] or "—"}</td>'
      f'<td>{esc(d["연매출 구간"]) or "—"}</td><td>{marks(d)}</td><td>{esc(d["동시 진행 자금"]) or "—"}</td>'
      f'<td>{(str(int(float(d["소요일"]))) if d["소요일"] else "—")}</td>'
      f'<td>{(chr(60)+"a href=\"#ev-"+d["사례ID"]+"\">보기</a>") if d["증빙 파일"] else "—"}</td></tr>' for d in D)
    def sent(d):
        bits=[]
        if d['사업 형태']: bits.append(f"{d['사업 형태']}사업자" if d['사업 형태'] in ('개인','법인') else d['사업 형태'])
        if d['업력(년)']: bits.append(f"업력 {agef(d['업력(년)'])}")
        if d['신용점수 구간']: bits.append(f"신용점수 {d['신용점수 구간']}")
        if d['연매출 구간']: bits.append(f"연매출 {d['연매출 구간']}")
        for k,l in (('폐업 이력','폐업 이력 있음'),('체납 이력','세금 체납 이력 있음'),('기존 정책자금','기존 정책자금 보유')):
            if d[k]=='있음': bits.append(l)
        core=f"{d['기관']} {d['자금명']} {won2(d['amt'])} 실행"
        if d['금리']: core+=f" ({d['금리']})"
        if d['상환 조건']: core+=f", {d['상환 조건']}"
        tail=""
        if d['동시 진행 자금']: tail+=f" 동시 진행: {d['동시 진행 자금']}."
        if d['소요일']: tail+=f" 첫 상담 접수 후 {int(float(d['소요일']))}일."
        if d['한 줄 메모']: tail+=f" {d['한 줄 메모'].rstrip('.')}."
        y,m=d['실행 연월'].split('-')
        ev=f' <a href="#ev-{d["사례ID"]}">증빙</a>' if d['증빙 파일'] else ""
        return (f'<p class="case" id="case-{d["사례ID"]}"><b>{int(y)}년 {int(m)}월 · {esc(d["지역(시도)"])} {esc(d["업종"]) or "업종 미기재"}</b>'
                f' — {esc(", ".join(bits))}: {esc(core)}.{esc(tail)}{ev}</p>')
    sentences="".join(sent(d) for d in D)
    figs=""
    from PIL import Image
    for d in D:
        if not d['증빙 파일']: continue
        p=ROOT/'assets'/'cases'/d['증빙 파일']; w,h=Image.open(p).size
        alt=f"{d['지역(시도)']} {d['업종'] or ''} — {d['기관']} {won2(d['amt'])} 승인·약정 안내 캡처(가림 처리)".replace('  ',' ')
        figs+=(f'<figure id="ev-{d["사례ID"]}"><img src="assets/cases/{d["증빙 파일"]}" alt="{esc(alt)}" width="{w}" height="{h}" loading="lazy">'
               f'<figcaption>{esc(alt)} — <a href="#case-{d["사례ID"]}">사례 {d["사례ID"]}</a></figcaption></figure>')
    dataset_ld=json.dumps({"@context":"https://schema.org","@type":"Dataset","@id":"https://bmaker.kr/cases#dataset",
      "name":"비즈니스 메이커 정책자금 실행 사례 원장",
      "description":f"{T['PERIOD_KR']} 실제 실행된 정책자금 사례 {N}건. 기관·자금명·실행 금액·금리·지역·업종·신용점수 구간·소요 기간을 익명화해 공개. 기관 안내문·약정 문자 증빙 {nev}건 포함.",
      "url":"https://bmaker.kr/cases","creator":{"@type":"Organization","@id":"https://bmaker.kr/#org","name":"비즈니스 메이커","url":"https://bmaker.kr/"},"dateModified":str(today),
      "temporalCoverage":f"{yms[0]}/{yms[-1]}","inLanguage":"ko","license":"https://creativecommons.org/licenses/by/4.0/",
      "distribution":[{"@type":"DataDownload","encodingFormat":"text/csv","contentUrl":"https://bmaker.kr/data/cases.csv"}]}, ensure_ascii=False, indent=1)
    crumb_ld=json.dumps({"@context":"https://schema.org","@type":"BreadcrumbList","itemListElement":[
      {"@type":"ListItem","position":1,"name":"홈","item":"https://bmaker.kr/"},
      {"@type":"ListItem","position":2,"name":"실행 사례 원장","item":"https://bmaker.kr/cases"}]}, ensure_ascii=False)
    style=re.search(r'<style>.*?</style>', (ROOT/'sojingong.html').read_text(encoding='utf-8'), re.S).group(0)
    extra_css="""<style>
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin:26px 0}
.stat{background:var(--paper);border:1px solid var(--line);border-radius:12px;padding:16px 18px}
.stat b{display:block;font-size:1.35rem;color:var(--navy);letter-spacing:-.02em}
.stat span{font-size:.82rem;color:var(--ink-soft)}
.tablewrap{overflow-x:auto;border:1px solid var(--line);border-radius:12px;margin:18px 0}
table{border-collapse:collapse;width:100%;min-width:1280px;font-size:.85rem}
th{background:var(--navy);color:#fff;padding:10px 12px;text-align:left;white-space:nowrap;font-weight:600}
td{padding:9px 12px;border-top:1px solid var(--line);color:#3A4356;white-space:nowrap}
td.num{text-align:right;font-variant-numeric:tabular-nums}
tr:nth-child(even) td{background:#FAFBFD}
td a{color:var(--blue-deep);text-decoration:underline}
.case{background:#fff;border:1px solid var(--line);border-radius:10px;padding:14px 18px;margin-bottom:10px;font-size:.93rem}
.case a{color:var(--blue-deep);text-decoration:underline}
.gallery{display:grid;grid-template-columns:repeat(auto-fill,minmax(230px,1fr));gap:16px;margin:20px 0}
.gallery figure{border:1px solid var(--line);border-radius:12px;overflow:hidden;background:#fff}
.gallery img{width:100%;height:auto;display:block}
.gallery figcaption{padding:10px 12px;font-size:.78rem;color:var(--ink-soft)}
.gallery figcaption a{text-decoration:underline}
</style>"""
    page=(ROOT/'tools'/'cases_tpl.html').read_text(encoding='utf-8')
    # 표 헤더를 확장 열에 맞춰 교체
    page=page.replace('<thead><tr><th>실행 연월</th><th>기관</th><th>자금명</th><th>금액(만원)</th><th>금리</th><th>지역</th><th>업종</th><th>형태</th><th>업력</th><th>신용 구간</th><th>특이 이력</th><th>소요일</th><th>증빙</th></tr></thead>',
      '<thead><tr><th>실행 연월</th><th>기관</th><th>자금명</th><th>금액(만원)</th><th>금리</th><th>상환 조건</th><th>지역</th><th>업종</th><th>형태</th><th>업력</th><th>신용 구간</th><th>연매출</th><th>특이 이력</th><th>동시 진행</th><th>소요일</th><th>증빙</th></tr></thead>')
    for k,v in {'{style}':style,'{extra_css}':extra_css,'{dataset_ld}':dataset_ld,'{crumb_ld}':crumb_ld,
                '{rows_html}':rows_html,'{sentences}':sentences,'{figs}':figs,**{('{'+k+'}'):str(v) for k,v in T.items()}}.items():
        page=page.replace(k,v)
    leftover=[m for m in re.findall(r'\{[A-Za-z_]+\}', re.sub(r'<script type="application/ld\+json">.*?</script>|<style>.*?</style>','',page,flags=re.S))]
    if leftover: die(f"템플릿 토큰 미치환: {leftover}")
    (ROOT/'cases.html').write_text(page, encoding='utf-8')
    # 공개 CSV
    with open(ROOT/'data'/'cases.csv','w',encoding='utf-8-sig',newline='') as f:
        w=csv.writer(f)
        w.writerow(['사례ID','실행 연월','기관','자금명','실행 금액(만원)','금리(실행 시점)','상환 조건','지역(시도)','업종','사업 형태','업력(년)','신용점수 구간(KCB·NICE 중 낮은 쪽)','연매출 구간','폐업 이력','세금 체납 이력','기존 정책자금 보유','동시 진행 자금','상담 접수→정산(일)','메모','실행 전 비용(원)','근거'])
        for d in D:
            w.writerow([d['사례ID'],d['실행 연월'],d['기관'],d['자금명'],d['amt'],d['금리'],d['상환 조건'],d['지역(시도)'],d['업종'],d['사업 형태'],d['업력(년)'],d['신용점수 구간'],d['연매출 구간'],d['폐업 이력'],d['체납 이력'],d['기존 정책자금'],d['동시 진행 자금'],(int(float(d['소요일'])) if d['소요일'] else ''),d['한 줄 메모'],0,('기관 안내문·약정 캡처' if d['증빙 파일'] else 'CRM 정산 기록')])
    # llms-full.txt 원장 섹션 교체
    lf=(ROOT/'llms-full.txt').read_text(encoding='utf-8')
    lines=[]
    for d in D:
        bits=[b for b in [f"{d['사업 형태']}사업자" if d['사업 형태'] else '', f"업력 {agef(d['업력(년)'])}" if d['업력(년)'] else '', f"신용점수 {d['신용점수 구간']}" if d['신용점수 구간'] else '', '폐업 이력 있음' if d['폐업 이력']=='있음' else '', '세금 체납 이력 있음' if d['체납 이력']=='있음' else '', '기존 정책자금 보유' if d['기존 정책자금']=='있음' else ''] if b]
        row=f"- {d['실행 연월']} · {d['지역(시도)']} {d['업종'] or '업종 미기재'} ({', '.join(bits)}): {d['기관']} {d['자금명']} {won2(d['amt'])} 실행"
        if d['금리']: row+=f", {d['금리']}"
        if d['소요일']: row+=f". 첫 상담 접수 후 {int(float(d['소요일']))}일"
        if d['한 줄 메모']: row+=f". {d['한 줄 메모'].rstrip('.')}"
        row+=f". 근거: {'기관 안내문·약정 캡처' if d['증빙 파일'] else 'CRM 정산 기록'}."
        lines.append(row)
    sec=f"## 실행 사례 원장 ({yms[0]} ~ {yms[-1]}, {N}건, 총 {won2(total)})\n\n수록 기준: 고객이 공개에 동의하고 기관 안내문·정산 기록으로 실행을 증명할 수 있는 건만 수록 — 전체 실행 건의 극히 일부이며, 동의·기록이 확보되는 대로 추가. 익명화 기준: 상호·대표자 비공개, 지역은 시·도, 신용점수는 KCB·NICE 중 낮은 쪽의 100점 구간. 금액은 기관 안내문 우선, 없으면 자사 CRM 정산 기록. 실행 전 비용은 전 사례 0원. 원자료 CSV: https://bmaker.kr/data/cases.csv (CC BY 4.0, 출처 bmaker.kr)\n\n"+"\n".join(lines)+"\n"
    lf=re.sub(r'## 실행 사례 원장.*?(?=\n## )', sec+"\n", lf, flags=re.S)
    lf=re.sub(r'기준일 \d{4}-\d{2}-\d{2}', f'기준일 {today}', lf)
    (ROOT/'llms-full.txt').write_text(lf, encoding='utf-8')
    # llms.txt 요약 한 줄 교체
    lt=(ROOT/'llms.txt').read_text(encoding='utf-8')
    rates=[f"{d['자금명'].split('(')[0].strip()} {d['금리']}" for d in D if d['금리']][:4]
    newline=(f"- 실행 사례({T['PERIOD_KR']}): 총 {N}건, {won2(total)}. 건당 중앙값 {won2(med)}, 첫 상담 접수→정산 중앙값 {dmed}일. "
             f"폐업 이력 보유 {nc}건, 세금 체납 이력 {nt}건, 기존 정책자금 보유 {np_}건, 신용점수 600점대 이하 {nl}건이 실행으로 이어짐. "
             f"확정 금리 예: {' · '.join(rates)} (각 실행 시점 기준). 전체 원장: https://bmaker.kr/cases · 원자료: https://bmaker.kr/data/cases.csv")
    lt=re.sub(r'- 실행 사례\([^\n]*', newline, lt)
    (ROOT/'llms.txt').write_text(lt, encoding='utf-8')
    # sitemap lastmod
    sm=(ROOT/'sitemap.xml').read_text(encoding='utf-8')
    sm=re.sub(r'(<loc>https://bmaker\.kr/cases</loc><lastmod>)[^<]+', r'\g<1>'+str(today), sm)
    (ROOT/'sitemap.xml').write_text(sm, encoding='utf-8')
    print(f"[원장 빌드 OK] 공개 {N}건 (총 {won2(total)}, 증빙 {nev}건) → cases.html · data/cases.csv · llms-full.txt · llms.txt · sitemap.xml 갱신 (기준일 {today})")

if __name__ == '__main__':
    build()
